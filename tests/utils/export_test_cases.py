#!/usr/bin/env python3
"""
Unified Test Case Exporter
Exports both Selenium and Unit test cases to Excel format
"""
import os
import ast
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Paths
PROJECT_ROOT = Path(__file__).parent.parent.parent
SELENIUM_DIR = PROJECT_ROOT / "tests" / "selenium"
UNIT_DIR = PROJECT_ROOT / "tests" / "unit"
OUTPUT_DIR = PROJECT_ROOT / "tests" / "selenium" / "reports"


def extract_selenium_tests(directory):
    """Extract test cases from Selenium test files."""
    test_cases = []

    for file_path in directory.rglob("*.py"):
        if file_path.name.startswith("conftest") or file_path.name.startswith("_"):
            continue

        with open(file_path, "r", encoding="utf-8") as f:
            try:
                tree = ast.parse(f.read(), filename=str(file_path))
            except SyntaxError:
                continue

        for node in ast.walk(tree):
            if isinstance(node, ast.FunctionDef) and node.name.startswith("test_"):
                docstring = ast.get_docstring(node) or ""
                test_cases.append({
                    "file": file_path.relative_to(directory),
                    "test_name": node.name,
                    "description": docstring.strip(),
                    "line": node.lineno
                })

    return test_cases


def extract_unit_tests(directory):
    """Extract test cases from unit test files."""
    test_cases = []

    for file_path in directory.rglob("*.py"):
        if file_path.name.startswith("conftest") or file_path.name.startswith("_"):
            continue

        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        for idx, line in enumerate(lines, start=1):
            if line.strip().startswith("def test_"):
                match = re.search(r"def\s+(test_[a-zA-Z0-9_]+)", line)
                if match:
                    test_name = match.group(1)

                    # Try to get docstring from next few lines
                    description = ""
                    if idx < len(lines) and '"""' in lines[idx]:
                        description = lines[idx].strip().strip('"""')

                    test_cases.append({
                        "file": file_path.relative_to(directory),
                        "test_name": test_name,
                        "description": description,
                        "line": idx
                    })

    return test_cases


def create_excel_report(selenium_tests, unit_tests, output_path):
    """Create Excel workbook with both test suites."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Header styling
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # ===== Selenium Tests Sheet =====
    ws_selenium = wb.create_sheet("Selenium Tests")
    ws_selenium.append(["File", "Test Name", "Description", "Line"])

    # Style header
    for cell in ws_selenium[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add data
    for test in sorted(selenium_tests, key=lambda x: str(x["file"])):
        ws_selenium.append([
            str(test["file"]),
            test["test_name"],
            test["description"],
            test["line"]
        ])

    # Adjust column widths
    ws_selenium.column_dimensions['A'].width = 30
    ws_selenium.column_dimensions['B'].width = 40
    ws_selenium.column_dimensions['C'].width = 60
    ws_selenium.column_dimensions['D'].width = 10

    # ===== Unit Tests Sheet =====
    ws_unit = wb.create_sheet("Unit Tests")
    ws_unit.append(["File", "Test Name", "Description", "Line"])

    # Style header
    for cell in ws_unit[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add data
    for test in sorted(unit_tests, key=lambda x: str(x["file"])):
        ws_unit.append([
            str(test["file"]),
            test["test_name"],
            test["description"],
            test["line"]
        ])

    # Adjust column widths
    ws_unit.column_dimensions['A'].width = 30
    ws_unit.column_dimensions['B'].width = 40
    ws_unit.column_dimensions['C'].width = 60
    ws_unit.column_dimensions['D'].width = 10

    # ===== Summary Sheet =====
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["Test Suite", "Total Tests"])
    ws_summary.append(["Selenium Tests", len(selenium_tests)])
    ws_summary.append(["Unit Tests", len(unit_tests)])
    ws_summary.append(["TOTAL", len(selenium_tests) + len(unit_tests)])

    # Style summary
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws_summary['A4'].font = Font(bold=True)
    ws_summary['B4'].font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15

    # Save
    wb.save(output_path)
    return len(selenium_tests), len(unit_tests)


def main():
    """Main execution."""
    print("=" * 60)
    print("Test Case Exporter - AI Cooking Assistant")
    print("=" * 60)

    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Extract tests
    print(f"\n📋 Extracting Selenium tests from: {SELENIUM_DIR}")
    selenium_tests = extract_selenium_tests(SELENIUM_DIR)
    print(f"   Found {len(selenium_tests)} Selenium tests")

    print(f"\n📋 Extracting Unit tests from: {UNIT_DIR}")
    unit_tests = extract_unit_tests(UNIT_DIR)
    print(f"   Found {len(unit_tests)} Unit tests")

    # Create report
    output_file = OUTPUT_DIR / "All_Test_Cases.xlsx"
    print(f"\n📊 Creating Excel report: {output_file}")

    sel_count, unit_count = create_excel_report(selenium_tests, unit_tests, output_file)

    print(f"\n✅ Export complete!")
    print(f"   - Selenium tests: {sel_count}")
    print(f"   - Unit tests: {unit_count}")
    print(f"   - Total: {sel_count + unit_count}")
    print(f"   - Location: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
